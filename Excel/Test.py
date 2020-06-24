import openpyxl

workbook = openpyxl.load_workbook("Employees.xlsx")

# Excel Properties
print (workbook.properties)
print (workbook.sheetnames)
print (workbook.active)

# Adding Sheet
sheet = workbook['EmployeeData']
workbook.create_sheet('TestSheet')
workbook.save("Employees.xlsx")

#Removing Sheet
sheet = workbook['TestSheet']
workbook.remove(sheet)
#del workbook['TestSheet']
workbook.save("Employees.xlsx")


