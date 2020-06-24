import openpyxl

workbook = openpyxl.load_workbook("Employees.xlsx")

# Sheet Operation1
sheet = workbook['EmployeeData']
#print (sheet.title)
#print (dir(sheet)
#print (sheet.active_cell)
#print (sheet.dimensions)
#print (sheet.sheet_format)
#print (sheet.sheet_properties)
#print (sheet.sheet_state)
#print (sheet.max_row)
#print (sheet.max_column)

# Sheet Operation2
#for i in sheet.values:
#    print (i)

# Sheet Operation3
print (sheet['B7'].value)
print (sheet.cell(row = 6, column = 2).value)

# Sheet Operation4
cell = sheet['B9']
print (cell.row)
print (cell.column)
print (cell.coordinate)
print (cell.data_type)
print (cell.encoding)

# Sheet Operation5
cell = sheet['B2']
cell.value = 'David'
workbook.save("Employees.xlsx")
print (cell.parent)

